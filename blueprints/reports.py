import io, csv, logging
from datetime import datetime, date, timedelta
import calendar
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, make_response, current_app
from flask_login import login_required, current_user
from app import db
from models import CleanWaterPlant, WaterTankLevel, WaterTank, CustomerReading, Customer
from sqlalchemy import func, extract, case
from utils import generate_daily_report, generate_monthly_report, check_permissions
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import unicodedata
import re

bp = Blueprint('reports', __name__)
logger = logging.getLogger(__name__)


@bp.route('/reports')
@login_required
def reports():
    if not check_permissions(current_user.role, ['accounting', 'plant_manager', 'leadership', 'admin']):
        flash('You do not have permission to access this page', 'error')
        return redirect(url_for('dashboard'))

    return render_template('reports.html')

def _xlsx_response(wb: Workbook, filename_base: str):
    """
    Serialize openpyxl Workbook to an in-memory XLSX and return send_file response.
    """
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"{filename_base}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def _build_sample_report_wb(report_type: str, start_dt: date, end_dt: date) -> Workbook:
    """
    TẠM THỜI: Tạo workbook mẫu có dữ liệu tối thiểu để đảm bảo mở được bằng Excel.
    Bạn có thể thay phần fill data bằng truy vấn thực tế.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'BaoCao'
    ws.append(['Báo cáo', report_type])
    ws.append(['Từ ngày', start_dt.strftime('%Y-%m-%d')])
    ws.append(['Đến ngày', end_dt.strftime('%Y-%m-%d')])
    ws.append([])

    # Header dữ liệu ví dụ
    ws.append(['Ngày', 'Mô tả', 'Giá trị'])
    cur = start_dt
    v = 1000
    while cur <= end_dt:
        ws.append([cur.strftime('%Y-%m-%d'), 'Dữ liệu mẫu', v])
        v += 37
        cur += timedelta(days=1)
    return wb

def _date_range(start_dt: date, end_dt: date):
    cur = start_dt
    while cur <= end_dt:
        yield cur
        cur += timedelta(days=1)

def _build_clean_water_plant_report_wb(start_dt: date, end_dt: date) -> Workbook:
    """Builds an Excel workbook for 'BÁO CÁO NHÀ MÁY NƯỚC SẠCH (m3)'.

    Mapping requirements:
    - NƯỚC CẤP              -> CleanWaterPlant.clean_water_output (per day)
    - NƯỚC THÔ JASAN        -> CleanWaterPlant.raw_water_jasan (per day)
    - LƯỢNG NƯỚC TẠI CÁC BỂ CHỨA (clean water tanks)
        Columns: BỂ 1200, BỂ 2000, BỂ 4000 -> WaterTankLevel.level for tanks with capacity 1200/2000/4000 (tank_type='clean_water')
    - NƯỚC SẠCH MỘT SỐ DOANH NGHIỆP (customer clean water reading)
        Columns: NHUỘM HY, LEEHING HT, LEEHING TT, JASAN, LỆ TINH
        Data: sum of CustomerReading.clean_water_reading per customer per day (0 if none)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'BÁO CÁO'

    # Styles
    header_fill = PatternFill(start_color='FFECD9', end_color='FFECD9', fill_type='solid')
    subheader_fill = PatternFill(start_color='FFF5E6', end_color='FFF5E6', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center')

    # Title row (merge across columns A to M)
    ws.merge_cells('A1:M1')
    ws['A1'] = 'BÁO CÁO NHÀ MÁY NƯỚC SẠCH (m3)'
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = center

    # Header rows
    # Row 2: group headers
    headers_row2 = [
        'STT', 'NGÀY', 'NƯỚC CẤP',
        'LƯỢNG NƯỚC TẠI CÁC BỂ CHỨA', '', '',
        'NƯỚC SẠCH MỘT SỐ DOANH NGHIỆP', '', '', '', '',
        'NƯỚC THÔ JASAN', 'GHI CHÚ'
    ]
    ws.append(headers_row2)
    # Row 3: sub-headers
    sub_headers_row3 = [
        '', '', '',
        'BỂ 1200', 'BỂ 2000', 'BỂ 4000',
        'NHUỘM HY', 'LEEHING HT', 'LEEHING TT', 'JASAN', 'LỆ TINH',
        '', ''
    ]
    ws.append(sub_headers_row3)

    # Merge group cells
    ws.merge_cells('A2:A3')  # STT
    ws.merge_cells('B2:B3')  # NGÀY
    ws.merge_cells('C2:C3')  # NƯỚC CẤP
    ws.merge_cells('D2:F2')  # BỂ CHỨA
    ws.merge_cells('G2:K2')  # DOANH NGHIỆP
    ws.merge_cells('L2:L3')  # NƯỚC THÔ JASAN
    ws.merge_cells('M2:M3')  # GHI CHÚ

    # Style header rows
    for row_idx in (2, 3):
        for col in range(1, 14):  # A..M (13 columns)
            c = ws.cell(row=row_idx, column=col)
            c.font = Font(bold=True)
            c.alignment = center
            c.fill = header_fill if row_idx == 2 else subheader_fill
            c.border = border

    # Column widths
    widths = [6, 12, 12, 10, 10, 10, 12, 12, 12, 12, 12, 14, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Data preparation
    dates = list(_date_range(start_dt, end_dt))

    # Clean water plant data
    plant_rows = db.session.query(CleanWaterPlant.date, CleanWaterPlant.clean_water_output, CleanWaterPlant.raw_water_jasan) \
        .filter(CleanWaterPlant.date >= start_dt, CleanWaterPlant.date <= end_dt) \
        .all()
    clean_map = {r[0]: float(r[1] or 0) for r in plant_rows}
    raw_jasan_map = {r[0]: float(r[2] or 0) for r in plant_rows}

    # Tanks: only clean water tanks
    tanks = WaterTank.query.filter(WaterTank.tank_type == 'clean_water').all()
    # Map tank_id to label (by capacity or name)
    def tank_label(t: WaterTank):
        cap = int((t.capacity or 0))
        if cap in (1200, 2000, 4000):
            return f'BỂ {cap}'
        # Fallback: try parse from name
        name = (t.name or '').lower()
        for c in (1200, 2000, 4000):
            if str(c) in name:
                return f'BỂ {c}'
        return t.name or f'TANK {t.id}'

    desired_tank_labels = ['BỂ 1200', 'BỂ 2000', 'BỂ 4000']
    tank_by_label = {}
    for t in tanks:
        lbl = tank_label(t)
        if lbl in desired_tank_labels and lbl not in tank_by_label:
            tank_by_label[lbl] = t.id

    # Fetch levels
    levels = db.session.query(WaterTankLevel.date, WaterTankLevel.tank_id, WaterTankLevel.level) \
        .filter(WaterTankLevel.date >= start_dt, WaterTankLevel.date <= end_dt, WaterTankLevel.tank_id.in_(tank_by_label.values() if tank_by_label else [-1])) \
        .all()
    levels_map = {lbl: {} for lbl in desired_tank_labels}
    for dt_val, tank_id, level in levels:
        # find label
        for lbl, tid in tank_by_label.items():
            if tid == tank_id:
                levels_map[lbl][dt_val] = float(level or 0)
                break

    # --- Map cột theo ID (bỏ nhận diện theo tên) ---
    customer_columns = ['NHUỘM HY', 'LEEHING HT', 'LEEHING TT', 'JASAN', 'LỆ TINH']
    cust_col_map = {
        'NHUỘM HY':   [13],
        'LEEHING HT': [25],   # nếu về sau tách TT riêng thì thêm id vào 'LEEHING TT'
        'LEEHING TT': [],
        'JASAN':      [21],
        'LỆ TINH':    [26],
    }

    # --- Hệ số theo ID: (k1, k2) ứng với (ΔĐH1, ΔĐH2) ---
    FACTOR_BY_ID = {
        13: (10.0, 1.0),  # Nhuộm HY
        21: (1.0,  1.0),  # Jasan
        25: (1.0, 10.0),  # Lee Hing HT
        26: (1.0,  1.0),  # Lệ Tinh
    }

    # ==== Tính Δ cho 2 đồng hồ (window function) ====
    r1 = func.coalesce(CustomerReading.clean_water_reading, 0.0)
    lag_r1 = func.lag(r1).over(
        partition_by=CustomerReading.customer_id,
        order_by=CustomerReading.date
    )
    delta1 = case(((r1 - func.coalesce(lag_r1, r1)) < 0, 0.0),
                else_=(r1 - func.coalesce(lag_r1, r1)))

    r2 = func.coalesce(getattr(CustomerReading, "clean_water_reading_2", 0.0), 0.0)
    lag_r2 = func.lag(r2).over(
        partition_by=CustomerReading.customer_id,
        order_by=CustomerReading.date
    )
    delta2 = case(((r2 - func.coalesce(lag_r2, r2)) < 0, 0.0),
                else_=(r2 - func.coalesce(lag_r2, r2)))

    # ==== CASE hệ số theo ID (fallback 1.0 nếu không match) ====
    whens_k1 = [(Customer.id == cid, f1) for cid, (f1, _) in FACTOR_BY_ID.items()]
    whens_k2 = [(Customer.id == cid, f2) for cid, (_, f2) in FACTOR_BY_ID.items()]
    k1 = case(*whens_k1, else_=1.0) if whens_k1 else 1.0
    k2 = case(*whens_k2, else_=1.0) if whens_k2 else 1.0
    clean_delta_expr = (delta1 * k1) + (delta2 * k2)

    # ==== Dải ngày cho LAG: cần (start_dt - 1) ====
    calc_start = start_dt - timedelta(days=1)

    # === LỚP 1: subquery tính delta theo dòng (dùng lag) ===
    delta_sq = (
        db.session.query(
            CustomerReading.date.label('date'),
            CustomerReading.customer_id.label('customer_id'),
            clean_delta_expr.label('clean_delta')
        )
        .join(Customer, Customer.id == CustomerReading.customer_id)
        .filter(
            CustomerReading.date >= calc_start,
            CustomerReading.date <= end_dt,
            Customer.is_active.is_(True)
        )
    ).subquery()

    # === LỚP 2: tổng theo (date, customer) (sum trên subquery, KHÔNG dùng lag ở đây) ===
    readings = (
        db.session.query(
            delta_sq.c.date,
            delta_sq.c.customer_id,
            func.sum(delta_sq.c.clean_delta).label('total_clean')
        )
        .group_by(delta_sq.c.date, delta_sq.c.customer_id)
        .all()
    )

    # --- Đưa vào 5 cột DN (bỏ ngày calc_start khi hiển thị) ---
    cust_series = {k: {} for k in customer_columns}
    for d, cid, total_clean in readings:
        if d < start_dt:
            continue
        val = float(total_clean or 0.0)
        for col, ids in cust_col_map.items():
            if cid in ids:
                cust_series[col][d] = float(cust_series[col].get(d, 0.0)) + val
                break

    # 2) Ghép vào các cột doanh nghiệp đã định nghĩa
    cust_series = {k: {} for k in customer_columns}
    for d, cid, total_clean in readings:
        total_val = float(total_clean or 0.0)
        # Đưa bản ghi (date,cid) về đúng cột (NHUỘM HY, LEEHING HT/TT, JASAN, LỆ TINH)
        for col, ids in cust_col_map.items():
            if cid in ids:
                # Nếu cùng ngày/cột đã có giá trị từ KH khác (cùng nhóm), cộng dồn
                cust_series[col][d] = float(cust_series[col].get(d, 0.0)) + total_val
                break

    # Build per-column map
    cust_series = {k: {} for k in customer_columns}
    for d, cid, total_clean in readings:
        for col, ids in cust_col_map.items():
            if cid in ids:
                cust_series[col][d] = float(total_clean or 0) + float(cust_series[col].get(d, 0))
                break

    # Write data rows
    row_idx = 4
    stt = 1
    for d in dates:
        ws.cell(row=row_idx, column=1, value=stt).alignment = center
        ws.cell(row=row_idx, column=2, value=d.strftime('%d/%m/%Y')).alignment = center

        # NƯỚC CẤP
        ws.cell(row=row_idx, column=3, value=clean_map.get(d, 0)).alignment = right

        # Tanks D,E,F
        ws.cell(row=row_idx, column=4, value=levels_map['BỂ 1200'].get(d, 0)).alignment = right
        ws.cell(row=row_idx, column=5, value=levels_map['BỂ 2000'].get(d, 0)).alignment = right
        ws.cell(row=row_idx, column=6, value=levels_map['BỂ 4000'].get(d, 0)).alignment = right

        # Customers G..K
        for i, col_name in enumerate(['NHUỘM HY', 'LEEHING HT', 'LEEHING TT', 'JASAN', 'LỆ TINH'], start=7):
            ws.cell(row=row_idx, column=i, value=cust_series[col_name].get(d, 0)).alignment = right

        # NƯỚC THÔ JASAN (L)
        ws.cell(row=row_idx, column=12, value=raw_jasan_map.get(d, 0)).alignment = right

        # GHI CHÚ (M) empty for now
        ws.cell(row=row_idx, column=13, value='')

        # Borders for the row
        for col in range(1, 14):
            ws.cell(row=row_idx, column=col).border = border
        row_idx += 1
        stt += 1

    # Apply number formatting with thousand separators
    for r in ws.iter_rows(min_row=4, min_col=3, max_col=13, max_row=row_idx-1):
        for c in r:
            c.number_format = '#,##0'

    # Freeze panes below headers
    ws.freeze_panes = 'A4'

    return wb


def _build_nmns_monthly_power_chem_wb(start_dt: date, end_dt: date) -> Workbook:
    """BÁO CÁO SỐ LIỆU ĐỊNH MỨC SỬ DỤNG ĐIỆN VÀ HOÁ CHẤT (NMNS)

    - Cột tháng: từ T01/<year> đến T<current>/<year> theo end_dt
    - Hàng nội dung (10 dòng):
        1. Số điện (kWh) = sum(electricity)
        2. Số điện/ 1m3 nước sạch ( kw/m3) = điện / Tổng nước sạch
        3. PAC (kg) = sum(pac_usage)
        4. PAC/m3 nước sạch (kg/m3) = PAC / Tổng nước sạch
        5. Xút (kg) = sum(naoh_usage)
        6. Xút/m3 nước sạch (kg/m3) = Xút / Tổng nước sạch
        7. Polymer (kg) = sum(polymer_usage)
        8. Polymer/m3 nước sạch (g/m3) = (Polymer / Tổng nước sạch) * 1000
        9. Tổng nước sạch (m3) = sum(clean_water_output)
       10. Lượng nước sạch TB ngày = Tổng nước sạch / số ngày trong tháng

    Ghi chú: tính trong phạm vi từ 01/01/<year> tới end_dt.
    """
    # Phạm vi từ đầu năm đến ngày kết thúc
    year = end_dt.year
    start_year_dt = date(year, 1, 1)
    last_month = end_dt.month

    # Lấy tổng theo tháng từ DB
    q = db.session.query(
        extract('year', CleanWaterPlant.date).label('y'),
        extract('month', CleanWaterPlant.date).label('m'),
        func.sum(func.coalesce(CleanWaterPlant.electricity, 0)).label('electricity'),
        func.sum(func.coalesce(CleanWaterPlant.pac_usage, 0)).label('pac'),
        func.sum(func.coalesce(CleanWaterPlant.naoh_usage, 0)).label('naoh'),
        func.sum(func.coalesce(CleanWaterPlant.polymer_usage, 0)).label('polymer'),
        func.sum(func.coalesce(CleanWaterPlant.clean_water_output, 0)).label('water')
    ).filter(
        CleanWaterPlant.date >= start_year_dt,
        CleanWaterPlant.date <= end_dt
    ).group_by('y', 'm').order_by('y', 'm')

    monthly = {(int(r.y), int(r.m)): {
        'electricity': float(r.electricity or 0),
        'pac': float(r.pac or 0),
        'naoh': float(r.naoh or 0),
        'polymer': float(r.polymer or 0),
        'water': float(r.water or 0),
    } for r in q.all()}

    # Workbook setup
    wb = Workbook()
    ws = wb.active
    ws.title = 'BÁO CÁO'

    # Styles
    header_fill = PatternFill(start_color='8E1340', end_color='8E1340', fill_type='solid')  # tím đậm gần với ảnh
    header_font = Font(color='FFFFFF', bold=True, size=12)
    table_header_fill = PatternFill(start_color='C24C86', end_color='C24C86', fill_type='solid')
    table_header_font = Font(color='FFFFFF', bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center')
    right = Alignment(horizontal='right', vertical='center')

    # Title rows
    total_cols = 2 + last_month  # B = Nội dung, cộng C.. for months
    last_col_letter = chr(64 + total_cols)
    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'] = 'BÁO CÁO SỐ LIỆU ĐỊNH MỨC SỬ DỤNG ĐIỆN VÀ HOÁ CHẤT'
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = center

    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'] = f'NHÀ MÁY NƯỚC SẠCH THÁNG {str(end_dt.month).zfill(2)}/{year}'
    ws['A2'].font = Font(size=12, bold=True)
    ws['A2'].alignment = center

    # Header row (TT | Nội dung | T01/YYYY ... TMM/YYYY)
    headers = ['TT', 'Nội dung'] + [f'T{str(m).zfill(2)}/{year}' for m in range(1, last_month + 1)]
    ws.append(headers)
    header_row = 3
    for idx in range(1, total_cols + 1):
        c = ws.cell(row=header_row, column=idx)
        c.fill = table_header_fill
        c.font = table_header_font
        c.alignment = center
        c.border = border

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 38
    for col_idx in range(3, total_cols + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 14

    # Rows definitions
    rows_def = [
        ('Số điện (kWh)', 'electricity', '#,##0'),
        ('Số điện/ 1m3 nước sạch ( kw/m3)', 'electricity_per_m3', '0.000'),
        ('PAC (kg)', 'pac', '#,##0'),
        ('PAC/m3 nước sạch (kg/m3)', 'pac_per_m3', '0.000'),
        ('Xút (kg)', 'naoh', '#,##0'),
        ('Xút/m3 nước sạch (kg/m3)', 'naoh_per_m3', '0.000'),
        ('Polymer (kg)', 'polymer', '#,##0'),
        ('Polymer/m3 nước sạch (g/m3)', 'polymer_per_m3_g', '0.000'),
        ('Tổng nước sạch (m3)', 'water', '#,##0'),
        ('Lượng nước sạch TB ngày', 'avg_water_per_day', '#,##0'),
    ]

    current_row = header_row + 1
    for i, (label, key, numfmt) in enumerate(rows_def, start=1):
        ws.cell(row=current_row, column=1, value=i).alignment = center
        ws.cell(row=current_row, column=2, value=label).alignment = left
        # fill monthly values
        for m in range(1, last_month + 1):
            data = monthly.get((year, m), {'electricity': 0, 'pac': 0, 'naoh': 0, 'polymer': 0, 'water': 0})
            water = data['water']
            value = 0
            if key == 'electricity':
                value = data['electricity']
            elif key == 'pac':
                value = data['pac']
            elif key == 'naoh':
                value = data['naoh']
            elif key == 'polymer':
                value = data['polymer']
            elif key == 'water':
                value = water
            elif key == 'electricity_per_m3':
                value = (data['electricity'] / water) if water else 0
            elif key == 'pac_per_m3':
                value = (data['pac'] / water) if water else 0
            elif key == 'naoh_per_m3':
                value = (data['naoh'] / water) if water else 0
            elif key == 'polymer_per_m3_g':
                value = ((data['polymer'] / water) * 1000) if water else 0
            elif key == 'avg_water_per_day':
                days = calendar.monthrange(year, m)[1]
                value = (water / days) if days else 0

            col_idx = 2 + m  # month columns start at C (3)
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.alignment = right
            cell.number_format = numfmt

        # borders for whole row
        for col_idx in range(1, total_cols + 1):
            ws.cell(row=current_row, column=col_idx).border = border

        current_row += 1

    # Apply borders for header too (already applied) and set freeze panes
    ws.freeze_panes = 'C4'

    # Footer signatures (simple approximation)
    footer_start = current_row + 2
    # Người lập
    ws.merge_cells(start_row=footer_start, start_column=2, end_row=footer_start, end_column=3)
    ws.cell(row=footer_start, column=2, value='NGƯỜI LẬP').alignment = center
    ws.cell(row=footer_start, column=2).font = Font(bold=True)

    # Phòng quản lý hạ tầng / Trưởng phòng
    mid_col_start = max(5, total_cols // 2)
    mid_col_end = min(total_cols, mid_col_start + 2)
    ws.merge_cells(start_row=footer_start, start_column=mid_col_start, end_row=footer_start, end_column=mid_col_end)
    ws.cell(row=footer_start, column=mid_col_start, value='PHÒNG QUẢN LÝ HẠ TẦNG').alignment = center
    ws.cell(row=footer_start, column=mid_col_start).font = Font(bold=True)

    ws.merge_cells(start_row=footer_start + 1, start_column=mid_col_start, end_row=footer_start + 1, end_column=mid_col_end)
    ws.cell(row=footer_start + 1, column=mid_col_start, value='TRƯỞNG PHÒNG').alignment = center

    # Names a few rows below
    name_row = footer_start + 6
    ws.merge_cells(start_row=name_row, start_column=2, end_row=name_row, end_column=3)
    ws.cell(row=name_row, column=2, value='CAO MINH HIẾU').alignment = center

    ws.merge_cells(start_row=name_row, start_column=mid_col_start, end_row=name_row, end_column=mid_col_end)
    ws.cell(row=name_row, column=mid_col_start, value='TÔ NGỌC CƯỜNG').alignment = center

    return wb


def _build_monthly_wastewater_1_wb(start_dt: date, end_dt: date) -> Workbook:
    """BÁO CÁO TỔNG HỢP NMNT SỐ 1 - theo tháng (T1..Tn + Tổng cộng)

    Chỉ tiêu theo yêu cầu:
      1. Nước thải theo BB chốt với DN (m3) = sum(clean_water_reading * customer.water_ratio) theo tháng
      2. Nước thải theo ĐH tại nhà máy (m3) = sum(WastewaterPlant.wastewater_meter) (plant 1)
      3. Nước thải theo Đầu vào TQT (m3) = sum(input_flow_tqt)
      4. Nước thải theo Đầu Ra TQT (m3) = sum(output_flow_tqt)
      5. Bùn (kg) = sum(sludge_output)
      6. Tỷ lệ bùn kg/m3 theo nước BB chốt = Bùn / (Nước thải BB DN)
      7. Điện (kW) = sum(electricity)
      8. Hóa chất sử dụng (kg) = sum(chemical_usage)
      9. Tỷ lệ điện (kW/m3) theo nước BB chốt = Điện / (Nước thải BB DN)
    """
    from models import WastewaterPlant, CustomerReading, Customer

    year = end_dt.year
    start_year_dt = date(year, 1, 1)
    last_month = end_dt.month

    # Build monthly buckets 1..last_month
    months = list(range(1, last_month + 1))

    # 1) BB DN = sum(wastewater_reading) + sum(wastewater_calculated) per month (logic mới)
    bb_rows = db.session.query(
        extract('year', CustomerReading.date).label('y'),
        extract('month', CustomerReading.date).label('m'),
        func.sum(func.coalesce(CustomerReading.wastewater_reading, 0)).label('ww_read'),
        func.sum(func.coalesce(CustomerReading.wastewater_calculated, 0)).label('ww_calc'),
    ).filter(
        # CustomerReading.customer_id == 1,
        CustomerReading.date >= start_year_dt,
        CustomerReading.date <= end_dt
    ).group_by('y', 'm').all()
    bb_by_month = {
        (int(r.y), int(r.m)): float((r.ww_read or 0) + (r.ww_calc or 0))
        for r in bb_rows
    }

    # 2..5,7,8 from WastewaterPlant for plant_number==1
    wp = db.session.query(
        extract('year', WastewaterPlant.date).label('y'),
        extract('month', WastewaterPlant.date).label('m'),
        func.sum(func.coalesce(WastewaterPlant.wastewater_meter, 0)).label('meter'),
        func.sum(func.coalesce(WastewaterPlant.input_flow_tqt, 0)).label('tqt_in'),
        func.sum(func.coalesce(WastewaterPlant.output_flow_tqt, 0)).label('tqt_out'),
        func.sum(func.coalesce(WastewaterPlant.sludge_output, 0)).label('sludge'),
        func.sum(func.coalesce(WastewaterPlant.electricity, 0)).label('electricity'),
        func.sum(func.coalesce(WastewaterPlant.chemical_usage, 0)).label('chem')
    ).filter(
        WastewaterPlant.date >= start_year_dt,
        WastewaterPlant.date <= end_dt,
        WastewaterPlant.plant_number == 1
    ).group_by('y', 'm').all()
    wp_by_month = {(int(r.y), int(r.m)): r for r in wp}

    # Prepare values per month
    def get_month_val(key, m):
        row = wp_by_month.get((year, m))
        if not row:
            return 0.0
        return float(getattr(row, key) or 0)

    bb_vals = [bb_by_month.get((year, m), 0.0) for m in months]
    meter_vals = [get_month_val('meter', m) for m in months]
    tqt_in_vals = [get_month_val('tqt_in', m) for m in months]
    tqt_out_vals = [get_month_val('tqt_out', m) for m in months]
    sludge_vals = [get_month_val('sludge', m) for m in months]
    electricity_vals = [get_month_val('electricity', m) for m in months]
    chem_vals = [get_month_val('chem', m) for m in months]

    # Ratios per month
    def safe_ratio(num, den):
        return (num / den) if den else 0.0
    sludge_ratio = [safe_ratio(sludge_vals[i], bb_vals[i]) for i in range(len(months))]
    elec_ratio = [safe_ratio(electricity_vals[i], bb_vals[i]) for i in range(len(months))]

    # Totals (sum) and overall ratios using total numerators/denominators
    total_bb = sum(bb_vals)
    total_meter = sum(meter_vals)
    total_tqt_in = sum(tqt_in_vals)
    total_tqt_out = sum(tqt_out_vals)
    total_sludge = sum(sludge_vals)
    total_electricity = sum(electricity_vals)
    total_chem = sum(chem_vals)
    total_sludge_ratio = safe_ratio(total_sludge, total_bb)
    total_elec_ratio = safe_ratio(total_electricity, total_bb)

    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'BÁO CÁO'

    header_font = Font(bold=True)
    title_center = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    total_cols = 2 + last_month + 1  # STT, Nội dung, T1..Tn, Tổng cộng
    last_col_letter = chr(64 + total_cols)

    # Title rows (approximation of screenshot)
    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'] = 'CÔNG TY CP PTHT DỆT MAY PHỐ NỐI'
    ws['A1'].alignment = title_center
    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'] = 'PHÒNG QUẢN LÝ HẠ TẦNG'
    ws['A2'].alignment = title_center
    ws['A2'].font = Font(bold=True)
    ws.merge_cells(f'A4:{last_col_letter}4')
    ws['A4'] = f'BÁO CÁO BÙN, ĐIỆN, NƯỚC NMXLNT  SỐ 1 NĂM {year}'
    ws['A4'].alignment = title_center
    ws['A4'].font = Font(size=12, bold=True)

    # Table header
    headers = ['Stt', 'Nội dung'] + [f'T{m}' for m in months] + ['Tổng cộng']
    ws.append([''] * 2)  # row 3 spacer
    ws.append(headers)   # row 5 (since row 1,2,3,4 already)
    header_row = 5
    for c in range(1, total_cols + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = header_font
        cell.alignment = title_center
        cell.border = border

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 48
    for i in range(3, total_cols + 1):
        ws.column_dimensions[chr(64 + i)].width = 14

    # Data rows definition
    rows = [
        (1, 'Nước thải theo BB chốt với DN (m3)', bb_vals, total_bb, '#,##0'),
        (2, 'Nước thải theo ĐH tại nhà máy (m3)', meter_vals, total_meter, '#,##0'),
        (3, 'Nước thải theo Đầu vào TQT (m3)', tqt_in_vals, total_tqt_in, '#,##0'),
        (4, 'Nước thải theo Đầu Ra TQT (m3)', tqt_out_vals, total_tqt_out, '#,##0'),
        (5, 'Bùn (kg)', sludge_vals, total_sludge, '#,##0'),
        (6, 'Tỷ lệ bùn kg/m3 theo nước BB chốt', sludge_ratio, total_sludge_ratio, '0.00'),
        (7, 'Điện (kw)', electricity_vals, total_electricity, '#,##0'),
        (8, 'Hóa chất sử dụng (kg)', chem_vals, total_chem, '#,##0'),
        (8, 'Tỷ lệ điện(kw/m3) theo nước BB chốt', elec_ratio, total_elec_ratio, '0.00'),
    ]

    r = header_row + 1
    red_font = Font(color='FF0000', bold=True)
    for (stt, label, series, total_value, numfmt) in rows:
        ws.cell(row=r, column=1, value=stt).alignment = title_center
        ws.cell(row=r, column=2, value=label)
        # Monthly values
        for i, mval in enumerate(series):
            col = 3 + i
            c = ws.cell(row=r, column=col, value=mval)
            c.number_format = numfmt
            c.alignment = Alignment(horizontal='right')
            c.border = border
        # Total column
        ctot = ws.cell(row=r, column=total_cols, value=total_value)
        ctot.number_format = numfmt
        ctot.alignment = Alignment(horizontal='right')
        ctot.font = red_font
        ctot.border = border
        # Borders for first two columns
        ws.cell(row=r, column=1).border = border
        ws.cell(row=r, column=2).border = border
        r += 1

    # Borders for header row already set; set for any empty cells in header
    for col in range(1, total_cols + 1):
        ws.cell(row=header_row, column=col).border = border

    ws.freeze_panes = ws['C6']

    return wb


def _build_monthly_wastewater_2_wb(start_dt: date, end_dt: date) -> Workbook:
    """BÁO CÁO TỔNG HỢP NMNT SỐ 2 - theo tháng (T1..Tn + Tổng cộng)

    Chỉ tiêu theo yêu cầu (same as NMNT 1 but for plant_number = 2):
      1. Nước thải theo BB chốt với DN (m3) = sum(wastewater_reading) + sum(wastewater_calculated) theo tháng
      2. Nước thải theo ĐH tại nhà máy (m3) = sum(WastewaterPlant.wastewater_meter) (plant 2)
      3. Nước thải theo Đầu vào TQT (m3) = sum(input_flow_tqt)
      4. Nước thải theo Đầu Ra TQT (m3) = sum(output_flow_tqt)
      5. Bùn (kg) = sum(sludge_output)
      6. Tỷ lệ bùn kg/m3 theo nước BB chốt = Bùn / (Nước thải BB DN)
      7. Điện (kW) = sum(electricity)
      8. Hóa chất sử dụng (kg) = sum(chemical_usage)
      9. Tỷ lệ điện (kW/m3) theo nước BB chốt = Điện / (Nước thải BB DN)
    """
    from models import WastewaterPlant, CustomerReading, Customer

    year = end_dt.year
    start_year_dt = date(year, 1, 1)
    last_month = end_dt.month

    # Build monthly buckets 1..last_month
    months = list(range(1, last_month + 1))

    # 1) BB DN = sum(wastewater_reading) + sum(wastewater_calculated) per month 
    bb_rows = db.session.query(
        extract('year', CustomerReading.date).label('y'),
        extract('month', CustomerReading.date).label('m'),
        func.sum(func.coalesce(CustomerReading.wastewater_reading, 0)).label('ww_read'),
        func.sum(func.coalesce(CustomerReading.wastewater_calculated, 0)).label('ww_calc'),
    ).filter(
        CustomerReading.date >= start_year_dt,
        CustomerReading.date <= end_dt
    ).group_by('y', 'm').all()
    bb_by_month = {
        (int(r.y), int(r.m)): float((r.ww_read or 0) + (r.ww_calc or 0))
        for r in bb_rows
    }

    # 2..5,7,8 from WastewaterPlant for plant_number==2
    wp = db.session.query(
        extract('year', WastewaterPlant.date).label('y'),
        extract('month', WastewaterPlant.date).label('m'),
        func.sum(func.coalesce(WastewaterPlant.wastewater_meter, 0)).label('meter'),
        func.sum(func.coalesce(WastewaterPlant.input_flow_tqt, 0)).label('tqt_in'),
        func.sum(func.coalesce(WastewaterPlant.output_flow_tqt, 0)).label('tqt_out'),
        func.sum(func.coalesce(WastewaterPlant.sludge_output, 0)).label('sludge'),
        func.sum(func.coalesce(WastewaterPlant.electricity, 0)).label('electricity'),
        func.sum(func.coalesce(WastewaterPlant.chemical_usage, 0)).label('chem')
    ).filter(
        WastewaterPlant.date >= start_year_dt,
        WastewaterPlant.date <= end_dt,
        WastewaterPlant.plant_number == 2
    ).group_by('y', 'm').all()
    wp_by_month = {(int(r.y), int(r.m)): r for r in wp}

    # Prepare values per month
    def get_month_val(key, m):
        row = wp_by_month.get((year, m))
        if not row:
            return 0.0
        return float(getattr(row, key) or 0)

    bb_vals = [bb_by_month.get((year, m), 0.0) for m in months]
    meter_vals = [get_month_val('meter', m) for m in months]
    tqt_in_vals = [get_month_val('tqt_in', m) for m in months]
    tqt_out_vals = [get_month_val('tqt_out', m) for m in months]
    sludge_vals = [get_month_val('sludge', m) for m in months]
    electricity_vals = [get_month_val('electricity', m) for m in months]
    chem_vals = [get_month_val('chem', m) for m in months]

    # Ratios per month
    def safe_ratio(num, den):
        return (num / den) if den else 0.0
    sludge_ratio = [safe_ratio(sludge_vals[i], bb_vals[i]) for i in range(len(months))]
    elec_ratio = [safe_ratio(electricity_vals[i], bb_vals[i]) for i in range(len(months))]

    # Totals (sum) and overall ratios using total numerators/denominators
    total_bb = sum(bb_vals)
    total_meter = sum(meter_vals)
    total_tqt_in = sum(tqt_in_vals)
    total_tqt_out = sum(tqt_out_vals)
    total_sludge = sum(sludge_vals)
    total_electricity = sum(electricity_vals)
    total_chem = sum(chem_vals)
    total_sludge_ratio = safe_ratio(total_sludge, total_bb)
    total_elec_ratio = safe_ratio(total_electricity, total_bb)

    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'BÁO CÁO'

    header_font = Font(bold=True)
    title_center = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    total_cols = 2 + last_month + 1  # STT, Nội dung, T1..Tn, Tổng cộng
    last_col_letter = chr(64 + total_cols)

    # Title rows (approximation of screenshot)
    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'] = 'CÔNG TY CP PTHT DỆT MAY PHỐ NỐI'
    ws['A1'].alignment = title_center
    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'] = 'PHÒNG QUẢN LÝ HẠ TẦNG'
    ws['A2'].alignment = title_center
    ws['A2'].font = Font(bold=True)
    ws.merge_cells(f'A4:{last_col_letter}4')
    ws['A4'] = f'BÁO CÁO BÙN, ĐIỆN, NƯỚC NMXLNT  SỐ 2 NĂM {year}'
    ws['A4'].alignment = title_center
    ws['A4'].font = Font(size=12, bold=True)

    # Table header
    headers = ['Stt', 'Nội dung'] + [f'T{m}' for m in months] + ['Tổng cộng']
    ws.append([''] * 2)  # row 3 spacer
    ws.append(headers)   # row 5 (since row 1,2,3,4 already)
    header_row = 5
    for c in range(1, total_cols + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = header_font
        cell.alignment = title_center
        cell.border = border

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 48
    for i in range(3, total_cols + 1):
        ws.column_dimensions[chr(64 + i)].width = 14

    # Data rows definition (including chemical_usage row 8 for plant 2)
    rows = [
        # (1, 'Nước thải theo BB chốt với DN (m3)', bb_vals, total_bb, '#,##0'),
        (1, 'Nước thải theo ĐH tại nhà máy (m3)', meter_vals, total_meter, '#,##0'),
        (2, 'Nước thải theo Đầu vào TQT (m3)', tqt_in_vals, total_tqt_in, '#,##0'),
        (3, 'Nước thải theo Đầu Ra TQT (m3)', tqt_out_vals, total_tqt_out, '#,##0'),
        (4, 'Bùn (kg)', sludge_vals, total_sludge, '#,##0'),
        (5, 'Tỷ lệ bùn kg/m3 theo nước BB chốt', sludge_ratio, total_sludge_ratio, '0.00'),
        (6, 'Điện (kw)', electricity_vals, total_electricity, '#,##0'),
        (7, 'Hóa chất sử dụng (kg)', chem_vals, total_chem, '#,##0'),
        (8, 'Tỷ lệ điện(kw/m3) theo nước BB chốt', elec_ratio, total_elec_ratio, '0.00'),
    ]

    r = header_row + 1
    red_font = Font(color='FF0000', bold=True)
    for (stt, label, series, total_value, numfmt) in rows:
        ws.cell(row=r, column=1, value=stt).alignment = title_center
        ws.cell(row=r, column=2, value=label)
        # Monthly values
        for i, mval in enumerate(series):
            col = 3 + i
            c = ws.cell(row=r, column=col, value=mval)
            c.number_format = numfmt
            c.alignment = Alignment(horizontal='right')
            c.border = border
        # Total column
        ctot = ws.cell(row=r, column=total_cols, value=total_value)
        ctot.number_format = numfmt
        ctot.alignment = Alignment(horizontal='right')
        ctot.font = red_font
        ctot.border = border
        # Borders for first two columns
        ws.cell(row=r, column=1).border = border
        ws.cell(row=r, column=2).border = border
        r += 1

    # Borders for header row already set; set for any empty cells in header
    for col in range(1, total_cols + 1):
        ws.cell(row=header_row, column=col).border = border

    ws.freeze_panes = ws['C6']

    return wb

@bp.route('/generate-report/<report_type>')
@login_required
def generate_report(report_type):
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        format_type = request.args.get('format', 'excel')  # excel or pdf

        # Parse ngày hợp lệ
        if start_date and end_date:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
        else:
            end_dt = date.today()
            start_dt = end_dt - timedelta(days=30)

        if format_type == 'excel':
            # Excel: nhánh theo report_type
            if report_type == 'clean_water_plant':
                wb = _build_clean_water_plant_report_wb(start_dt, end_dt)
            elif report_type == 'monthly_clean_water':
                # Luôn lấy dữ liệu từ đầu năm tới tháng hiện tại
                wb = _build_nmns_monthly_power_chem_wb(date(end_dt.year, 1, 1), end_dt)
            elif report_type == 'monthly_wastewater_1':
                wb = _build_monthly_wastewater_1_wb(date(end_dt.year, 1, 1), end_dt)
            elif report_type == 'monthly_wastewater_2':
                wb = _build_monthly_wastewater_2_wb(date(end_dt.year, 1, 1), end_dt)
            else:
                wb = _build_sample_report_wb(report_type, start_dt, end_dt)
            filename = f"{report_type}_{date.today().strftime('%Y%m%d')}"
            return _xlsx_response(wb, filename)

        # Các định dạng khác (ví dụ pdf) vẫn dùng utils nếu bạn đã có sẵn
        if report_type == 'daily_clean_water':
            return generate_daily_report(start_date, end_date, format_type)
        else:
            return generate_monthly_report(report_type, start_date, end_date, format_type)

    except Exception as e:
        logger.exception("Error generating report")
        flash(f'Error generating report: {str(e)}', 'error')
        return redirect(url_for('reports'))
    

@bp.route('/reports/export', methods=['GET'])
@login_required
def export_reports():
    # TODO: thay data bằng dữ liệu thực tế của bạn
    headers = ['Ngày', 'Giếng', 'Sản lượng (m³)']
    rows = [
        ['2025-10-01', 'W1', 1234],
        ['2025-10-01', 'W2', 1456],
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = 'BaoCao'
    ws.append(headers)
    for r in rows:
        ws.append(r)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)  # QUAN TRỌNG

    filename = f"bao_cao_{date.today().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,  # thay cho attachment_filename
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@bp.route('/reports/export-csv', methods=['GET'])
@login_required
def export_reports_csv():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Ngày', 'Giếng', 'Sản lượng (m³)'])
    writer.writerow(['2025-10-01', 'W1', 1234])

    data = output.getvalue()
    resp = make_response(data.encode('utf-8-sig'))  # BOM để Excel mở đúng UTF-8
    resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
    resp.headers['Content-Disposition'] = 'attachment; filename=bao_cao.csv'
    return resp